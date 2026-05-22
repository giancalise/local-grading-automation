"""
grade_midterm.py
----------------
Batch grading for multi-problem SOLIDWORKS midterm exams.
Outputs a single Excel workbook with two sheets:

  Sheet 1 "Overview"  — Student x Problem grid, color-coded
  Sheet 2 "Detail"    — Full breakdown per student per problem

Usage
-----
    python grade_midterm.py ^
        --students  "C:\\path\\to\\Midterm_Folder" ^
        --solutions "C:\\path\\to\\Solutions_Folder" ^
        --output    "C:\\path\\to\\output" ^
        --assignment "MT26"

Folder structure expected
-------------------------
    students/
        Problem_1/   Baik.Ellen-MT26-1.SLDPRT ...
        Problem_2/   Baik.Ellen-MT26-2.SLDPRT ...
        ...
    solutions/
        SE00002301.SLDPRT   (Problem 1)
        SE00002302.SLDPRT   (Problem 2)
        ...

Rubric (edit below)
-------------------
"""

import argparse
import itertools
import os
import re
import shutil
import tempfile
import time
from pathlib import Path

import pythoncom
pythoncom.CoInitialize()

# ---------------------------------------------------------------------------
# Rubric
# ---------------------------------------------------------------------------
WEIGHT_SHAPE     = 0.65
WEIGHT_VOLUME    = 0.10
WEIGHT_MATERIAL  = 0.10
WEIGHT_SKETCHES  = 0.15
VOLUME_TOLERANCE = 0.01   # ±1%
SHAPE_THRESHOLD  = 0.95
VOXEL_RES        = 24

# ---------------------------------------------------------------------------
# Imports
# ---------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from popup_dismisser import ensure_dismisser_running
from sw_connection import get_connection, recover_from_stall
from tool_export import export_file
from tool_metadata import _read_summary_properties
from tool_mass import _read_mass_properties, _read_material, get_mass_properties
from tool_sketch import _read_sketch_statuses
from tool_compare import _normalize_mesh, _is_valid_mesh, _voxel_iou
import trimesh
import numpy as np


# ---------------------------------------------------------------------------
# Excel styles
# ---------------------------------------------------------------------------
GREEN_FILL  = PatternFill("solid", fgColor="92D050")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL    = PatternFill("solid", fgColor="FF0000")
BLUE_FILL   = PatternFill("solid", fgColor="2E75B6")
GREY_FILL   = PatternFill("solid", fgColor="D9D9D9")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

def hdr_font():  return Font(bold=True, size=11, name="Arial", color="FFFFFFFF")
def subhdr_font(): return Font(bold=True, size=10, name="Arial")
def norm_font(): return Font(size=10, name="Arial")
def flag_font(): return Font(bold=True, size=10, name="Arial")
def red_font():  return Font(bold=True, size=10, name="Arial", color="FFFFFFFF")

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center")
WRP = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def extract_username(filename: str) -> str:
    return Path(filename).stem.split('-')[0].strip()


def find_solution_file(solutions_folder: str, problem_num: int):
    sol_folder = Path(solutions_folder)
    target = f"SE{2300 + problem_num:08d}.SLDPRT"
    candidate = sol_folder / target
    if candidate.exists():
        return candidate
    for f in sol_folder.glob("*.SLDPRT"):
        if re.search(rf'0*{problem_num}$', f.stem):
            return f
    return None


def find_problem_folders(students_folder: str):
    base = Path(students_folder)
    problems = []
    for d in base.iterdir():
        if not d.is_dir():
            continue
        m = re.search(r'(\d+)$', d.name)
        if m:
            problems.append((int(m.group(1)), d))
    return sorted(problems, key=lambda x: x[0])


def compute_grade(shape_score, volume_ok, material_ok, sketch_ok):
    shape_credit = (1.0 if shape_score >= SHAPE_THRESHOLD else shape_score) \
                   if volume_ok else shape_score
    shape_pts    = round(shape_credit * WEIGHT_SHAPE * 100, 1)
    volume_pts   = round((1.0 if volume_ok else 0.0) * WEIGHT_VOLUME * 100, 1)
    material_pts = round((1.0 if material_ok else 0.0) * WEIGHT_MATERIAL * 100, 1)
    sketch_pts   = round((1.0 if sketch_ok else 0.0) * WEIGHT_SKETCHES * 100, 1)
    return {
        "shape":    shape_pts,
        "volume":   volume_pts,
        "material": material_pts,
        "sketches": sketch_pts,
        "total":    round(shape_pts + volume_pts + material_pts + sketch_pts, 1),
    }


def grade_one_student(conn, student_path, student_stl,
                      sol_norm, solution_volume, solution_material):
    r = {
        "filename":              student_path.name,
        "username":              extract_username(student_path.name),
        "sw_author":             None,
        "mass_kg":               None,
        "volume_mm3":            None,
        "material":              None,
        "underdefined_sketches": [],
        "shape_score":           None,
        "volume_ok":             None,
        "material_ok":           None,
        "sketches_ok":           None,
        "grade":                 None,
        "plagiarism_flag":       False,
        "plagiarism_with":       "",
        "error":                 None,
    }

    try:
        doc, _ = conn.open_part_silent(str(student_path))
        doc.ForceRebuild3(False)

        meta = {"author": None, "last_saved_date": None,
                "custom_properties": {}, "raw_identity_properties": {}, "error": None}
        _read_summary_properties(doc, meta)
        r["sw_author"] = meta.get("author")

        mass = {"mass": None, "volume": None, "surface_area": None,
                "center_of_mass": None, "density": None,
                "material_assigned": False, "material_name": None, "error": None}
        _read_mass_properties(doc, mass)
        _read_material(doc, mass)
        r["mass_kg"]    = mass.get("mass")
        r["volume_mm3"] = mass.get("volume")
        r["material"]   = mass.get("material_name")

        sketches = _read_sketch_statuses(doc)
        r["underdefined_sketches"] = [s["name"] for s in sketches
                                      if s["status"] == "UNDERDEFINED"]

        stl_result = export_file(str(student_path), "STL", student_stl)
        conn.close_doc(str(student_path))

        if not stl_result["success"]:
            r["error"] = f"STL export: {stl_result['error']}"

    except Exception as e:
        r["error"] = str(e)
        try: conn.close_doc(str(student_path))
        except: pass
        if recover_from_stall(timeout_s=10.0):
            print("    ⚠ SW recovered")
        return r

    # Shape comparison
    if os.path.exists(student_stl) and not r["error"]:
        try:
            mesh = trimesh.load(student_stl, force="mesh")
            if _is_valid_mesh(mesh):
                norm = _normalize_mesh(mesh)
                best = 0.0
                for flips in itertools.product([1, -1], repeat=3):
                    try:
                        b = norm.copy()
                        b.vertices *= np.array(flips)
                        iou = _voxel_iou(sol_norm, b, VOXEL_RES)
                        if iou > best: best = iou
                    except: pass
                r["shape_score"] = round(best, 4)
        except Exception as e:
            r["error"] = f"shape: {e}"

    vol = r["volume_mm3"]
    r["volume_ok"]   = (bool(solution_volume and vol) and
                        abs(vol - solution_volume) / solution_volume <= VOLUME_TOLERANCE)
    r["material_ok"] = (bool(r["material"] and solution_material) and
                        r["material"].lower() == solution_material.lower())
    r["sketches_ok"] = len(r["underdefined_sketches"]) == 0
    r["grade"] = compute_grade(
        min(r["shape_score"] or 0.0, 1.0),
        r["volume_ok"], r["material_ok"], r["sketches_ok"]
    )

    import gc; gc.collect()
    return r


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(output_folder, assignment, problem_nums, summary_data):
    wb = Workbook()
    students = sorted(summary_data.keys())

    # ================================================================
    # Sheet 1: Overview
    # ================================================================
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.freeze_panes = "B3"

    last_col = get_column_letter(1 + len(problem_nums))

    # Title banner
    ws1.merge_cells(f"A1:{last_col}1")
    ws1["A1"] = (f"{assignment}  —  Grading Overview     "
                 f"✓ = 100%    ⚠ = Review needed    ⚑ = Plagiarism")
    ws1["A1"].font = hdr_font()
    ws1["A1"].fill = BLUE_FILL
    ws1["A1"].alignment = LFT
    ws1.row_dimensions[1].height = 18

    # Column headers
    ws1.cell(2, 1, "Student").font = subhdr_font()
    ws1.cell(2, 1).fill = GREY_FILL
    ws1.cell(2, 1).alignment = LFT
    ws1.cell(2, 1).border = thin_border()
    ws1.column_dimensions["A"].width = 22

    for j, pnum in enumerate(problem_nums, start=2):
        c = ws1.cell(2, j, f"P{pnum}")
        c.font = subhdr_font()
        c.fill = GREY_FILL
        c.alignment = CTR
        c.border = thin_border()
        ws1.column_dimensions[get_column_letter(j)].width = 7
    ws1.row_dimensions[2].height = 16

    # Student rows
    for i, username in enumerate(students, start=3):
        ws1.cell(i, 1, username).font = norm_font()
        ws1.cell(i, 1).alignment = LFT
        ws1.cell(i, 1).border = thin_border()

        for j, pnum in enumerate(problem_nums, start=2):
            rec = summary_data.get(username, {}).get(pnum)
            c = ws1.cell(i, j)
            c.alignment = CTR
            c.border = thin_border()

            if rec is None or rec.get("error"):
                c.value = "—"
                c.fill = GREY_FILL
                c.font = norm_font()
            elif rec.get("plagiarism_flag"):
                c.value = "⚑"
                c.fill = RED_FILL
                c.font = red_font()
            elif (rec.get("grade") or {}).get("total", 0) >= 100:
                c.value = "✓"
                c.fill = GREEN_FILL
                c.font = norm_font()
            else:
                c.value = "⚠"
                c.fill = YELLOW_FILL
                c.font = flag_font()

        ws1.row_dimensions[i].height = 15

    # Legend
    legend_row = len(students) + 4
    ws1.cell(legend_row, 1, "Legend:").font = subhdr_font()
    items = [
        ("✓", "Green  — 100%, no issues",         GREEN_FILL,  norm_font()),
        ("⚠", "Yellow — < 100%, review needed",   YELLOW_FILL, flag_font()),
        ("⚑", "Red    — Plagiarism flag",          RED_FILL,    red_font()),
        ("—", "Grey   — Not submitted / error",    GREY_FILL,   norm_font()),
    ]
    for k, (sym, desc, fl, fn) in enumerate(items):
        r = legend_row + 1 + k
        ws1.cell(r, 1, sym).fill = fl
        ws1.cell(r, 1).font = fn
        ws1.cell(r, 1).alignment = CTR
        ws1.cell(r, 1).border = thin_border()
        ws1.cell(r, 2, desc).font = norm_font()

    # ================================================================
    # Sheet 2: Detail
    # ================================================================
    ws2 = wb.create_sheet("Detail")
    ws2.freeze_panes = "A3"

    HEADERS = [
        "Student", "Problem", "Grade /100",
        f"Shape ({int(WEIGHT_SHAPE*100)}%)",
        f"Volume ({int(WEIGHT_VOLUME*100)}%)",
        f"Material ({int(WEIGHT_MATERIAL*100)}%)",
        f"Sketches ({int(WEIGHT_SKETCHES*100)}%)",
        "Underdefined Sketch Names",
        "Plagiarism", "SW Author",
    ]
    WIDTHS = [22, 10, 11, 12, 14, 14, 14, 28, 24, 14]

    # Title
    ws2.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws2["A1"] = f"{assignment}  —  Grading Detail"
    ws2["A1"].font = hdr_font()
    ws2["A1"].fill = BLUE_FILL
    ws2["A1"].alignment = LFT
    ws2.row_dimensions[1].height = 18

    # Headers
    for j, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws2.cell(2, j, h)
        c.font = subhdr_font()
        c.fill = GREY_FILL
        c.alignment = CTR
        c.border = thin_border()
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.row_dimensions[2].height = 30

    def tick(val):
        if val is True:  return "✓"
        if val is False: return "✗"
        return "?"

    # Data rows
    row = 3
    for username in students:
        for pnum in problem_nums:
            rec = summary_data[username].get(pnum)
            if rec is None:
                continue

            grade = (rec.get("grade") or {}).get("total")
            plag  = rec.get("plagiarism_flag", False)
            ud    = "; ".join(rec.get("underdefined_sketches", [])) or ""

            vals = [
                username,
                f"P{pnum}",
                grade if grade is not None else (rec.get("error") or "?"),
                rec.get("shape_score"),
                tick(rec.get("volume_ok")),
                tick(rec.get("material_ok")),
                tick(rec.get("sketches_ok")),
                ud,
                rec.get("plagiarism_with", "") if plag else "",
                rec.get("sw_author", ""),
            ]

            if rec.get("error"):
                row_fill = RED_FILL
            elif plag:
                row_fill = RED_FILL
            elif grade is not None and grade < 100:
                row_fill = YELLOW_FILL
            else:
                row_fill = WHITE_FILL

            for j, val in enumerate(vals, start=1):
                c = ws2.cell(row, j, val)
                c.font = norm_font()
                c.fill = row_fill
                c.border = thin_border()
                c.alignment = WRP if j == 8 else (LFT if j == 1 else CTR)

            ws2.row_dimensions[row].height = 15
            row += 1

    xlsx_path = Path(output_folder) / f"{assignment}_grades.xlsx"
    wb.save(str(xlsx_path))
    return xlsx_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def grade_midterm(students_folder, solutions_folder, output_folder, assignment):
    ensure_dismisser_running()
    conn = get_connection()

    Path(output_folder).mkdir(parents=True, exist_ok=True)
    tmp_dir = tempfile.mkdtemp(prefix="grading_mt_")
    t_total = time.monotonic()

    problem_folders = find_problem_folders(students_folder)
    print(f"\nFound {len(problem_folders)} problem folders")

    summary_data: dict[str, dict] = {}  # {username: {pnum: record}}
    author_map:   dict[str, list] = {}  # {sw_author: [usernames]}

    for problem_num, problem_folder in problem_folders:
        print(f"\n{'='*55}")
        print(f"PROBLEM {problem_num}: {problem_folder.name}")
        print(f"{'='*55}")

        sol_path = find_solution_file(solutions_folder, problem_num)
        if sol_path is None:
            print(f"  ⚠ No solution found for Problem {problem_num} — skipping")
            continue
        print(f"  Solution: {sol_path.name}")

        sol_stl = os.path.join(tmp_dir, f"sol_{problem_num}.stl")
        if not export_file(str(sol_path), "STL", sol_stl)["success"]:
            print(f"  ✗ Solution STL export failed — skipping")
            continue

        sol_props = get_mass_properties(str(sol_path))
        sol_vol   = sol_props.get("volume")
        sol_mat   = sol_props.get("material_name")
        sol_mesh  = trimesh.load(sol_stl, force="mesh")
        sol_norm  = _normalize_mesh(sol_mesh)
        print(f"  Volume: {sol_vol:.2f} mm³  Material: {sol_mat}")

        student_files = sorted(problem_folder.glob("*.SLDPRT"))
        # case-insensitive dedup
        seen = set()
        student_files = [f for f in student_files
                         if not (f.name.lower() in seen or seen.add(f.name.lower()))]
        print(f"  Students: {len(student_files)}\n")

        for i, sp in enumerate(student_files):
            t_s = time.monotonic()
            stu_stl = os.path.join(tmp_dir, f"p{problem_num}_s{i}.stl")
            rec = grade_one_student(conn, sp, stu_stl, sol_norm, sol_vol, sol_mat)

            username = rec["username"]
            if username not in summary_data:
                summary_data[username] = {}
            summary_data[username][problem_num] = rec

            # Track authors for plagiarism
            auth = rec.get("sw_author")
            if auth:
                if auth not in author_map:
                    author_map[auth] = []
                if username not in author_map[auth]:
                    author_map[auth].append(username)

            grade   = (rec.get("grade") or {}).get("total", "N/A")
            ud      = rec.get("underdefined_sketches", [])
            elapsed = round(time.monotonic() - t_s, 1)
            print(f"  [{i+1}/{len(student_files)}] {username:<22} "
                  f"grade={grade}/100  shape={rec.get('shape_score')}  "
                  f"vol={'✓' if rec.get('volume_ok') else '✗'}  "
                  f"sketches={'✓' if not ud else ud}  ({elapsed}s)")

    # Plagiarism flags
    print(f"\n{'='*55}\nPlagiarism analysis\n{'='*55}")
    flagged_count = 0
    for auth, users in author_map.items():
        if len(users) > 1:
            flagged_count += len(users)
            print(f"  ⚠ Author '{auth}' shared by: {', '.join(users)}")
            for username in users:
                others = [u for u in users if u != username]
                for pnum in summary_data.get(username, {}):
                    rec = summary_data[username][pnum]
                    if rec.get("sw_author") == auth:
                        rec["plagiarism_flag"] = True
                        rec["plagiarism_with"] = ", ".join(others)
    if flagged_count == 0:
        print("  No plagiarism flags.")

    # Write Excel
    print(f"\n{'='*55}\nWriting Excel report\n{'='*55}")
    problem_nums = sorted({p for p, _ in problem_folders})
    xlsx_path = write_excel(output_folder, assignment, problem_nums, summary_data)
    print(f"  → {xlsx_path}")

    total_time = round(time.monotonic() - t_total)
    print(f"\n{'='*55}")
    print(f"COMPLETE — {len(summary_data)} students, "
          f"{len(problem_nums)} problems, "
          f"{total_time}s (~{total_time//60}min)")
    print(f"{'='*55}")

    shutil.rmtree(tmp_dir, ignore_errors=True)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Grade a multi-problem SOLIDWORKS midterm exam"
    )
    parser.add_argument("--students",   required=True,
                        help="Parent folder with Problem_1, Problem_2 ... subfolders")
    parser.add_argument("--solutions",  required=True,
                        help="Folder with SE00002301.SLDPRT ... SE00002316.SLDPRT")
    parser.add_argument("--output",     required=True,
                        help="Output folder for the Excel report")
    parser.add_argument("--assignment", default="Midterm",
                        help="Assignment name (default: Midterm)")
    args = parser.parse_args()

    grade_midterm(args.students, args.solutions, args.output, args.assignment)
